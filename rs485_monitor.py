"""
RS-485 / USB Serial Monitor
---------------------------
A lightweight Windows desktop tool to connect to an RS-485-to-USB adapter,
show incoming hex-formatted lines live in a table, plot selected fields on
a live graph, and save the raw log to .txt or .xlsx.

Build into a standalone .exe with PyInstaller (see README.txt).
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import os
import sys
import json
import threading
import queue
from datetime import datetime

import serial
import serial.tools.list_ports

import openpyxl
from openpyxl.utils import get_column_letter

import matplotlib
matplotlib.use("TkAgg")
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.ticker import FuncFormatter

APP_TITLE = "RS-485 Serial Monitor"
APP_VERSION = "0.14"   # bump this each time I hand you a new file
RELEASE_DATE = "2026/07/02"
MAX_TABLE_ROWS = 4000            # cap what's ON SCREEN; saved data is never trimmed
GRAPH_REFRESH_MS = 300           # how often the graph redraws
QUEUE_POLL_MS = 50               # how often we drain incoming serial data
SAMPLE_INTERVAL_S = 0.02         # one line = 0.02s (device's fixed cycle, ~50/sec)

# Where saved tests live - next to the .exe when packaged (not PyInstaller's
# temp extraction folder), or next to the script when run directly.
if getattr(sys, "frozen", False):
    _APP_DIR = os.path.dirname(sys.executable)
else:
    _APP_DIR = os.path.dirname(os.path.abspath(__file__))
SAVED_TESTS_DIR = os.path.join(_APP_DIR, "saved_tests")

# Named meaning of each semicolon-separated field, in order (A..L in your
# Excel breakdown).
FIELD_NAMES = [
    "Encoder 1",         # A
    "Encoder 2",          # B
    "Effekt kW M1",       # C
    "Effekt kW M2",        # D
    "Str\u00f6m A M1",         # E
    "Str\u00f6m A M2",          # F
    "C211 M1",            # G
    "C211 M2",             # H
    "C212 M1",            # I
    "C212 M2",             # J
    "Port B",              # K
    "Felkod",               # L
]

# Short codes for the first six fields - used as the LIVE LOG TABLE header
# (replacing the full name there) and appended alongside the full name in
# the graph's checkbox labels. Fields without a code keep their full name
# everywhere.
FIELD_CODES = {
    0: "L111",   # Encoder 1
    1: "L121",   # Encoder 2
    2: "C231",   # Effekt kW M1
    3: "C241",   # Effekt kW M2
    4: "C251",   # Strom A M1
    5: "C261",   # Strom A M2
}

# Display-batch options: how many lines accumulate before the table repaints.
# Real data rate is ~50 lines/sec; these keep the UI from redrawing 50x/sec.
BATCH_OPTIONS = [
    ("0.2s  (flush every 10 lines)", 10),
    ("15s  (flush every 50 lines)", 50),
    ("30s  (flush every 100 lines)", 100),
    ("1 min  (flush every 250 lines)", 250),
    ("5 min+  (flush every 1000 lines)", 1000),
]

# Movement direction indicator, based on Encoder 1's direction of travel
# (increasing = opening, decreasing = closing). "None" shows nothing;
# "Show direction" shades sustained movement (above the filter duration
# set in Settings); "+ reversing" also flags quick direction flips
# shorter than that duration instead of just hiding them.
DIRECTION_MODES = ["None", "Show direction", "Show direction + reversing"]


def field_label(index):
    """Canonical full name for field position `index` (0-based). Used
    internally for unit/scale lookups - never changes based on display."""
    if index < len(FIELD_NAMES):
        return FIELD_NAMES[index]
    return f"Field {index + 1}"


def field_table_header(index):
    """What the Live Log table column header shows: the short code if one
    exists, otherwise the full name."""
    return FIELD_CODES.get(index, field_label(index))


def field_graph_label(index):
    """What the Live Graph checkbox shows: full name + code if one exists,
    otherwise just the full name."""
    code = FIELD_CODES.get(index)
    name = field_label(index)
    return f"{name} ({code})" if code else name


def field_scale_info(name):
    """Returns (scale_factor, unit, cap) for a given field name, used for
    the GRAPH and for the parenthetical value shown in the table. Raw hex
    values saved to file are always untouched.

    - Effekt kW fields: raw value is in units of 0.01 kW -> divide by 100,
      capped at 1.99 kW (spikes above that are almost certainly noise/glitch
      reads, not real readings).
    - Strom A (current) fields: raw value is in units of 0.1 A -> divide by
      10, capped at 20 A.
    - Encoder fields: raw value IS the degree value directly (1 count = 1
      degree, confirmed against real readings), no scaling needed.
    - Everything else: shown as-is, no unit, no cap.
    """
    if "kW" in name:
        return 0.01, "kW", 1.99
    if name.startswith("Str\u00f6m"):
        return 0.1, "A", 20.0
    if name.startswith("Encoder"):
        return 1.0, "\u00b0", None
    return 1.0, None, None


def carry_forward_fields(fields, last_known):
    """Given a raw fields list (blanks and all) and a mutable last_known
    list, returns a new list where blank entries are replaced with the
    last known non-blank value for that position. Mutates last_known in
    place. Used for display/graph only - raw data is never altered."""
    display_fields = []
    for i, f in enumerate(fields):
        if f == "" and i < len(last_known) and last_known[i] != "":
            display_fields.append(last_known[i])
        else:
            display_fields.append(f)
            if f != "" and i < len(last_known):
                last_known[i] = f
    return display_fields


def compute_scaled_value(index, hex_str):
    """Returns the scaled real-world value for a raw hex field string, or
    None if it can't be parsed as hex.

    - kW/A fields: capped at their ceiling (spikes clipped, not wrapped).
    - Degree fields: wrapped with modulo 360 so the result always falls in
      0-359, regardless of how high the raw counter goes.
    """
    scale, unit, cap = field_scale_info(field_label(index))
    try:
        raw = int(hex_str, 16)
    except ValueError:
        return None
    val = raw * scale
    if unit == "\u00b0":
        val = val % 360.0
    elif cap is not None and val > cap:
        val = cap
    return val


def build_buffers_from_lines(lines, field_count):
    """Given a list of raw semicolon-separated hex lines (e.g. from a saved
    test), replays carry-forward and scaling exactly like live data does,
    and returns (time_buffer, field_buffers) ready to plot. Elapsed time
    always starts at 0, so a saved test's shape lines up with live data
    regardless of when it was actually recorded."""
    last_known = [""] * field_count
    time_buffer = []
    field_buffers = [[] for _ in range(field_count)]
    for idx, line in enumerate(lines):
        fields = [f.strip() for f in line.split(";")]
        if len(fields) < field_count:
            fields = fields + [""] * (field_count - len(fields))
        elif len(fields) > field_count:
            fields = fields[:field_count]
        display_fields = carry_forward_fields(fields, last_known)
        time_buffer.append(round(idx * SAMPLE_INTERVAL_S, 2))
        for i, f in enumerate(display_fields):
            val = compute_scaled_value(i, f)
            field_buffers[i].append(val if val is not None else float("nan"))
    return time_buffer, field_buffers


def format_table_cell(index, hex_str):
    """Builds the table cell text: raw hex, plus the scaled real-world
    value in parentheses for fields that have a known unit."""
    _, unit, _ = field_scale_info(field_label(index))
    if unit is None:
        return hex_str
    val = compute_scaled_value(index, hex_str)
    if val is None:
        return hex_str
    val_str = f"{val:.2f}".replace(".", ",")
    if unit == "\u00b0":
        return f"{hex_str} ({val_str}{unit})"
    return f"{hex_str} ({val_str})"


class SerialMonitorApp:
    def __init__(self, root):
        self.root = root
        self.root.title(f"{APP_TITLE}  v{APP_VERSION}")
        self.root.geometry("1250x720")

        self.ser = None
        self.read_thread = None
        self.running = False
        self.data_queue = queue.Queue()

        # Full history of received lines: (timestamp, raw_line, [fields]).
        # This is NEVER trimmed - it's the source of truth for Save.
        self.raw_rows = []

        # Graph state - keeps full history too; redraw is downsampled for
        # performance, not the underlying data.
        self.field_count = 12
        self.field_vars = []
        self.field_buffers = []
        self.time_buffer = []
        self.sample_index = 0
        self.graph_paused = False

        # Some devices only transmit a field once and omit it (blank) on
        # later lines if it hasn't changed. We carry forward the last known
        # value for DISPLAY/GRAPH purposes only - the saved raw file always
        # stores exactly what was received, blank or not.
        self.last_known = [""] * self.field_count

        # Table batching state
        self.pending_rows = []     # rows waiting to be flushed to the table
        self.batch_size = 10
        self.batch_var = tk.StringVar(value=BATCH_OPTIONS[0][0])

        # Movement direction indicator (Encoder 1-based, factory 2.0s filter)
        self.direction_mode_var = tk.StringVar(value=DIRECTION_MODES[0])
        self.direction_filter_var = tk.StringVar(value="2,0")

        # Auto-save state
        self.autosave_path = None
        self.autosave_file = None

        # Saved-test overlays currently loaded onto the graph
        self.overlay_tests = []

        self._build_menu()
        self._build_ui()
        self._refresh_ports()
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)
        self.root.after(QUEUE_POLL_MS, self._poll_queue)
        self.root.after(GRAPH_REFRESH_MS, self._update_graph)

    # --------------------------------------------------------------- menu
    def _build_menu(self):
        menubar = tk.Menu(self.root)
        settings_menu = tk.Menu(menubar, tearoff=0)
        settings_menu.add_command(label="Settings...", command=self._open_settings_dialog)
        menubar.add_cascade(label="Settings", menu=settings_menu)
        self.root.config(menu=menubar)

    def _open_settings_dialog(self):
        win = tk.Toplevel(self.root)
        win.title("Settings")
        win.resizable(False, False)
        win.transient(self.root)

        frame = ttk.Frame(win, padding=14)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="Display Batching", font=("Segoe UI", 11, "bold")).pack(anchor="w")
        ttk.Label(frame,
                  text="Real data rate is ~50 lines/sec. The Live Log table\n"
                       "repaints once this many lines have arrived - nothing\n"
                       "is ever discarded, this only controls how often the\n"
                       "screen redraws.",
                  font=("Segoe UI", 8), foreground="#555", justify="left").pack(anchor="w", pady=(2, 8))

        batch_combo = ttk.Combobox(frame, textvariable=self.batch_var, width=32, state="readonly",
                                    values=[b[0] for b in BATCH_OPTIONS])
        batch_combo.pack(anchor="w", fill="x")
        batch_combo.bind("<<ComboboxSelected>>", self._on_batch_change)

        ttk.Separator(frame, orient="horizontal").pack(fill="x", pady=12)

        ttk.Label(frame, text="MDI (Movement Direction Indicator)", font=("Segoe UI", 11, "bold")).pack(anchor="w")
        ttk.Label(frame,
                  text="Based on Encoder 1's direction of travel. Movements\n"
                       "shorter than this are filtered out (or flagged as\n"
                       "\"reversing\" if that mode is selected on the Live\n"
                       "Graph tab) instead of shown as normal open/close.",
                  font=("Segoe UI", 8), foreground="#555", justify="left").pack(anchor="w", pady=(2, 8))
        filt_row = ttk.Frame(frame)
        filt_row.pack(anchor="w", fill="x")
        ttk.Label(filt_row, text="Minimum movement duration:").pack(side="left")
        ttk.Entry(filt_row, textvariable=self.direction_filter_var, width=6).pack(side="left", padx=(6, 4))
        ttk.Label(filt_row, text="seconds (factory: 2,0)").pack(side="left")

        ttk.Button(frame, text="Close", command=win.destroy).pack(anchor="e", pady=(14, 0))

        win.grab_set()

    # ------------------------------------------------------------------ UI
    def _build_ui(self):
        # ---- Left settings panel ----
        settings = ttk.Frame(self.root, padding=10)
        settings.pack(side="left", fill="y")

        ttk.Label(settings, text="Connection Settings", font=("Segoe UI", 11, "bold")).grid(
            row=0, column=0, columnspan=2, sticky="w", pady=(0, 8))

        r = 1
        ttk.Label(settings, text="Port:").grid(row=r, column=0, sticky="w")
        self.port_var = tk.StringVar()
        self.port_combo = ttk.Combobox(settings, textvariable=self.port_var, width=14, state="readonly")
        self.port_combo.grid(row=r, column=1, sticky="w")
        r += 1
        ttk.Button(settings, text="Refresh Ports", command=self._refresh_ports).grid(
            row=r, column=0, columnspan=2, sticky="we", pady=(2, 8))
        r += 1

        ttk.Label(settings, text="Baud Rate:").grid(row=r, column=0, sticky="w")
        self.baud_var = tk.StringVar(value="125000")
        ttk.Combobox(settings, textvariable=self.baud_var, width=14,
                     values=["9600", "19200", "38400", "57600", "115200", "125000", "230400", "460800"]).grid(
            row=r, column=1, sticky="w")
        r += 1

        ttk.Label(settings, text="Data Bits:").grid(row=r, column=0, sticky="w")
        self.databits_var = tk.StringVar(value="8")
        ttk.Combobox(settings, textvariable=self.databits_var, width=14, state="readonly",
                     values=["5", "6", "7", "8"]).grid(row=r, column=1, sticky="w")
        r += 1

        ttk.Label(settings, text="Parity:").grid(row=r, column=0, sticky="w")
        self.parity_var = tk.StringVar(value="None")
        ttk.Combobox(settings, textvariable=self.parity_var, width=14, state="readonly",
                     values=["None", "Even", "Odd", "Mark", "Space"]).grid(row=r, column=1, sticky="w")
        r += 1

        ttk.Label(settings, text="Stop Bits:").grid(row=r, column=0, sticky="w")
        self.stopbits_var = tk.StringVar(value="1")
        ttk.Combobox(settings, textvariable=self.stopbits_var, width=14, state="readonly",
                     values=["1", "1.5", "2"]).grid(row=r, column=1, sticky="w")
        r += 1

        self.rts_var = tk.BooleanVar(value=True)
        self.dtr_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(settings, text="RTS", variable=self.rts_var, command=self._apply_line_state).grid(
            row=r, column=0, sticky="w", pady=(6, 0))
        ttk.Checkbutton(settings, text="DTR", variable=self.dtr_var, command=self._apply_line_state).grid(
            row=r, column=1, sticky="w", pady=(6, 0))
        r += 1

        self.open_btn = ttk.Button(settings, text="Open Port", command=self._toggle_port)
        self.open_btn.grid(row=r, column=0, columnspan=2, sticky="we", pady=(10, 4))
        r += 1

        self.status_var = tk.StringVar(value="Closed")
        ttk.Label(settings, textvariable=self.status_var, foreground="red").grid(
            row=r, column=0, columnspan=2, sticky="w")
        r += 1

        ttk.Separator(settings, orient="horizontal").grid(row=r, column=0, columnspan=2, sticky="we", pady=10)
        r += 1

        ttk.Label(settings, text="Log / Save", font=("Segoe UI", 11, "bold")).grid(
            row=r, column=0, columnspan=2, sticky="w")
        r += 1

        ttk.Button(settings, text="Save as .txt", command=self._save_txt).grid(
            row=r, column=0, columnspan=2, sticky="we", pady=2)
        r += 1
        ttk.Button(settings, text="Save as .xlsx (raw hex)", command=self._save_xlsx).grid(
            row=r, column=0, columnspan=2, sticky="we", pady=2)
        r += 1
        ttk.Button(settings, text="Save as .xlsx (hex + values)", command=self._save_xlsx_with_values).grid(
            row=r, column=0, columnspan=2, sticky="we", pady=2)
        r += 1
        ttk.Button(settings, text="Export Graph...", command=self._export_graph).grid(
            row=r, column=0, columnspan=2, sticky="we", pady=2)
        r += 1
        ttk.Button(settings, text="Export All... (txt+xlsx\u00d72+graph)", command=self._export_all).grid(
            row=r, column=0, columnspan=2, sticky="we", pady=(2, 8))
        r += 1
        ttk.Button(settings, text="Save Current Test...", command=self._save_current_test).grid(
            row=r, column=0, columnspan=2, sticky="we", pady=2)
        r += 1
        ttk.Button(settings, text="Clear All", command=self._clear_all).grid(
            row=r, column=0, columnspan=2, sticky="we", pady=2)
        r += 1

        self.autosave_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(settings, text="Auto-save to .txt while running",
                         variable=self.autosave_var, command=self._toggle_autosave).grid(
            row=r, column=0, columnspan=2, sticky="w", pady=(8, 0))
        r += 1

        self.count_var = tk.StringVar(value="Received: 0")
        ttk.Label(settings, textvariable=self.count_var).grid(row=r, column=0, columnspan=2, sticky="w", pady=(10, 0))

        # ---- Right side: tabs ----
        notebook = ttk.Notebook(self.root)
        notebook.pack(side="right", fill="both", expand=True, padx=(0, 10), pady=10)

        self._build_log_tab(notebook)
        self._build_graph_tab(notebook)

    # --------------------------------------------------------------- Log tab
    def _build_log_tab(self, notebook):
        log_tab = ttk.Frame(notebook)
        notebook.add(log_tab, text="Live Log")

        top_row = ttk.Frame(log_tab)
        top_row.pack(anchor="w", fill="x", padx=5, pady=(5, 2))
        self.autoscroll_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(top_row, text="Auto-scroll", variable=self.autoscroll_var).pack(side="left")

        self.hex_view_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(top_row, text="Hex view (raw only)", variable=self.hex_view_var,
                         command=self._on_view_mode_change).pack(side="left", padx=(14, 0))

        # Packed before table_frame (which expands to fill the rest) so
        # this reserves its own thin row at the very bottom.
        ttk.Label(log_tab, text=f"v{APP_VERSION} \u2014 {RELEASE_DATE}",
                  font=("Segoe UI", 8), foreground="#888").pack(side="bottom", anchor="w", padx=6, pady=(0, 4))

        # Tighter row height / smaller font via a dedicated style
        style = ttk.Style()
        style.configure("Tight.Treeview", font=("Consolas", 9), rowheight=18)
        style.configure("Tight.Treeview.Heading", font=("Consolas", 9, "bold"))

        table_frame = ttk.Frame(log_tab)
        table_frame.pack(fill="both", expand=True, padx=5, pady=(0, 5))
        self.log_tree = ttk.Treeview(table_frame, show="headings", style="Tight.Treeview")
        yscroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.log_tree.yview)
        self.log_tree.configure(yscrollcommand=yscroll.set)
        self.log_tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)

        self.log_tree.tag_configure("fault", foreground="#c0392b", background="#fdecea")

        self._rebuild_table_columns()

    def _estimate_column_width(self, i):
        """Rough 'just wide enough' width for an Expanded-view column,
        based on the widest text that field is likely to show."""
        _, unit, _ = field_scale_info(field_label(i))
        if unit == "kW":
            content_len = len("0000 (1,99)")
        elif unit == "A":
            content_len = len("00 (20,0)")
        elif unit == "\u00b0":
            content_len = len("0167 (359,00\u00b0)")
        else:
            content_len = 6   # plain hex, e.g. "00".."FFFF"
        header_len = len(field_table_header(i))
        chars = max(content_len, header_len)
        return chars * 8 + 14

    def _rebuild_table_columns(self):
        if self.hex_view_var.get():
            self.log_tree.configure(columns=["raw"])
            self.log_tree.column("raw", width=700, anchor="w", stretch=True)
            self.log_tree.heading("raw", text="Raw Data")
            return

        cols = ["elapsed"] + [f"f{i}" for i in range(self.field_count)]
        self.log_tree.configure(columns=cols)
        self.log_tree.column("elapsed", width=60, anchor="e", stretch=False)
        self.log_tree.heading("elapsed", text="Time (s)")
        for i in range(self.field_count):
            key = f"f{i}"
            self.log_tree.column(key, width=self._estimate_column_width(i), anchor="center", stretch=False)
            self.log_tree.heading(key, text=field_table_header(i))

    def _on_view_mode_change(self):
        self._rebuild_table_columns()
        self._repopulate_table_from_history()

    def _repopulate_table_from_history(self):
        """Rebuilds the visible table from raw_rows in the current view
        mode. Replays carry-forward from the start so Expanded view still
        shows correct held-over values, even though only the last
        MAX_TABLE_ROWS end up on screen."""
        for item in self.log_tree.get_children():
            self.log_tree.delete(item)

        hex_mode = self.hex_view_var.get()
        last_known = [""] * self.field_count
        built = []
        for idx, (ts, line, fields) in enumerate(self.raw_rows):
            elapsed = round(idx * SAMPLE_INTERVAL_S, 2)
            display_fields = carry_forward_fields(fields, last_known)
            fault = False
            if display_fields:
                try:
                    fault = int(display_fields[-1], 16) != 0
                except ValueError:
                    fault = False
            built.append((elapsed, line, display_fields, fault))

        for elapsed, line, display_fields, fault in built[-MAX_TABLE_ROWS:]:
            if hex_mode:
                values = [line]
            else:
                values = [f"{elapsed:.2f}"] + [format_table_cell(i, f) for i, f in enumerate(display_fields)]
            tags = ("fault",) if fault else ()
            self.log_tree.insert("", "end", values=values, tags=tags)

        if self.autoscroll_var.get():
            children = self.log_tree.get_children()
            if children:
                self.log_tree.see(children[-1])

    # ------------------------------------------------------------- Graph tab
    def _build_graph_tab(self, notebook):
        graph_tab = ttk.Frame(notebook)
        notebook.add(graph_tab, text="Live Graph")

        # Packed first with side="bottom" so it reserves its own thin row
        # at the very bottom, regardless of what else expands above it.
        ttk.Label(graph_tab, text=f"v{APP_VERSION} \u2014 {RELEASE_DATE}",
                  font=("Segoe UI", 8), foreground="#888").pack(side="bottom", anchor="w", padx=6, pady=(0, 4))

        controls = ttk.Frame(graph_tab)
        controls.pack(side="top", fill="x", padx=5, pady=5)
        ttk.Label(controls, text="Plot field(s):").pack(side="left")

        self.field_checks_frame = ttk.Frame(controls)
        self.field_checks_frame.pack(side="left", padx=10)
        self._build_field_checkboxes()

        self.pause_btn = ttk.Button(controls, text="Pause Graph", command=self._toggle_pause)
        self.pause_btn.pack(side="right", padx=5)
        ttk.Button(controls, text="Reset Zoom", command=self._reset_zoom).pack(side="right", padx=5)
        ttk.Button(controls, text="Clear Graph", command=self._clear_graph).pack(side="right", padx=5)

        # ---- Threshold lines (optional, dashed) ----
        # Belastningsvakt (load guard) and Motorskydd (motor protection) are
        # per-motor, per-direction (open/close). Personskydd stays global.
        # Each control only shows up when its matching field is ticked
        # above - see _update_threshold_visibility().
        self.belastningsvakt_enabled = {"M1": tk.BooleanVar(value=False), "M2": tk.BooleanVar(value=False)}
        self.belastningsvakt_open = {"M1": tk.StringVar(value="0,70"), "M2": tk.StringVar(value="0,70")}
        self.belastningsvakt_close = {"M1": tk.StringVar(value="0,70"), "M2": tk.StringVar(value="0,70")}

        self.motorskydd_enabled = {"M1": tk.BooleanVar(value=False), "M2": tk.BooleanVar(value=False)}
        self.motorskydd_open = {"M1": tk.StringVar(value="0,8"), "M2": tk.StringVar(value="0,8")}
        self.motorskydd_close = {"M1": tk.StringVar(value="0,8"), "M2": tk.StringVar(value="0,8")}

        self.personskydd_var = tk.BooleanVar(value=False)
        self.personskydd_val = tk.StringVar(value="0,20")

        threshold_row = ttk.Frame(graph_tab)
        threshold_row.pack(side="top", fill="x", padx=5, pady=(0, 5))
        self.threshold_row = threshold_row

        def build_open_close_frame(parent, title, enabled_var, open_var, close_var, unit_text):
            frame = ttk.Frame(parent)
            ttk.Checkbutton(frame, text=title, variable=enabled_var, command=self._redraw_now).pack(side="left")
            ttk.Label(frame, text="\u00d6ppna:").pack(side="left", padx=(4, 0))
            ttk.Entry(frame, textvariable=open_var, width=5).pack(side="left")
            ttk.Label(frame, text="St\u00e4ng:").pack(side="left", padx=(4, 0))
            ttk.Entry(frame, textvariable=close_var, width=5).pack(side="left")
            ttk.Label(frame, text=unit_text).pack(side="left", padx=(2, 0))
            return frame

        self.belast_frame_m1 = build_open_close_frame(
            threshold_row, "Belastningsvakt M1", self.belastningsvakt_enabled["M1"],
            self.belastningsvakt_open["M1"], self.belastningsvakt_close["M1"], "kW")
        self.belast_frame_m2 = build_open_close_frame(
            threshold_row, "Belastningsvakt M2", self.belastningsvakt_enabled["M2"],
            self.belastningsvakt_open["M2"], self.belastningsvakt_close["M2"], "kW")

        self.personskydd_frame = ttk.Frame(threshold_row)
        ttk.Checkbutton(self.personskydd_frame, text="Personskydd", variable=self.personskydd_var,
                         command=self._redraw_now).pack(side="left")
        ttk.Entry(self.personskydd_frame, textvariable=self.personskydd_val, width=6).pack(side="left", padx=(2, 4))
        ttk.Label(self.personskydd_frame, text="kW (+0.25 / -0.10)").pack(side="left")

        self.motorskydd_frame_m1 = build_open_close_frame(
            threshold_row, "Motorskydd M1", self.motorskydd_enabled["M1"],
            self.motorskydd_open["M1"], self.motorskydd_close["M1"], "A (\u00b150%)")
        self.motorskydd_frame_m2 = build_open_close_frame(
            threshold_row, "Motorskydd M2", self.motorskydd_enabled["M2"],
            self.motorskydd_open["M2"], self.motorskydd_close["M2"], "A (\u00b150%)")

        # ---- Saved test overlays + MDI ----
        overlay_row = ttk.Frame(graph_tab)
        overlay_row.pack(side="top", fill="x", padx=5, pady=(0, 5))
        ttk.Label(overlay_row, text="Saved Tests:", font=("Segoe UI", 9, "bold")).pack(side="left", padx=(0, 6))
        self.saved_tests_var = tk.StringVar()
        self.saved_tests_combo = ttk.Combobox(overlay_row, textvariable=self.saved_tests_var, width=28,
                                               state="readonly", values=[])
        self.saved_tests_combo.pack(side="left", padx=(0, 6))
        ttk.Button(overlay_row, text="Overlay Selected", command=self._overlay_selected_test).pack(side="left", padx=(0, 6))
        ttk.Button(overlay_row, text="Clear Overlays", command=self._clear_overlays).pack(side="left", padx=(0, 6))
        ttk.Button(overlay_row, text="Import Hex...", command=self._import_hex_file).pack(side="left", padx=(0, 16))
        self._refresh_saved_tests_list()

        ttk.Label(overlay_row, text="MDI:", font=("Segoe UI", 9, "bold")).pack(side="left", padx=(0, 6))
        direction_combo = ttk.Combobox(overlay_row, textvariable=self.direction_mode_var, width=24,
                                        state="readonly", values=DIRECTION_MODES)
        direction_combo.pack(side="left")
        direction_combo.bind("<<ComboboxSelected>>", lambda e: self._redraw_now())

        self.fig = Figure(figsize=(6, 4), dpi=100)
        self.ax = self.fig.add_subplot(111)
        self.extra_axes = []
        self._user_zoomed = False
        self._zoom_reset_pending = False
        self._programmatic_xlim_update = False
        self.canvas = FigureCanvasTkAgg(self.fig, master=graph_tab)

        # Pan/zoom toolbar - drag to pan, the magnifier tool draws a zoom
        # rectangle. The graph always auto-scales to show everything by
        # default; it only starts preserving your view across live
        # redraws once you've actually used pan/zoom yourself (tracked via
        # the xlim_changed callback below), and "Reset Zoom" (or the
        # toolbar's home button) goes back to auto-scaling. Packed before
        # the canvas so it sits in its own fixed-height row above the
        # graph, which then expands to fill the rest.
        toolbar_frame = ttk.Frame(graph_tab)
        toolbar_frame.pack(side="top", fill="x", padx=5)
        self.nav_toolbar = NavigationToolbar2Tk(self.canvas, toolbar_frame)
        self.nav_toolbar.update()
        self.ax.callbacks.connect("xlim_changed", self._on_xlim_changed)

        self.canvas.get_tk_widget().pack(side="top", fill="both", expand=True, padx=5, pady=(0, 5))

        self.ax.set_xlabel("Time (s)")
        self.ax.grid(True, linewidth=0.3, alpha=0.6)

        self._update_threshold_visibility()

    def _build_field_checkboxes(self):
        for w in self.field_checks_frame.winfo_children():
            w.destroy()
        self.field_vars = []
        self.field_buffers = []
        for i in range(self.field_count):
            var = tk.BooleanVar(value=(i == 2))  # Effekt kW M1 ticked by default
            # Felkod is a fault code, not a measurement - keep it in the
            # table but don't offer it as a graph trace.
            if field_label(i) != "Felkod":
                cb = ttk.Checkbutton(self.field_checks_frame, text=field_graph_label(i), variable=var,
                                      command=self._redraw_now)
                cb.pack(side="left", padx=(0, 4))
            self.field_vars.append(var)
            self.field_buffers.append([])

    # ------------------------------------------------------------- helpers
    def _refresh_ports(self):
        ports = [p.device for p in serial.tools.list_ports.comports()]
        self.port_combo["values"] = ports
        if ports and not self.port_var.get():
            self.port_var.set(ports[0])

    def _parity_code(self):
        return {"None": serial.PARITY_NONE, "Even": serial.PARITY_EVEN,
                "Odd": serial.PARITY_ODD, "Mark": serial.PARITY_MARK,
                "Space": serial.PARITY_SPACE}[self.parity_var.get()]

    def _stopbits_code(self):
        return {"1": serial.STOPBITS_ONE, "1.5": serial.STOPBITS_ONE_POINT_FIVE,
                "2": serial.STOPBITS_TWO}[self.stopbits_var.get()]

    def _apply_line_state(self):
        if self.ser and self.ser.is_open:
            try:
                self.ser.rts = self.rts_var.get()
                self.ser.dtr = self.dtr_var.get()
            except Exception:
                pass

    def _on_batch_change(self, event=None):
        for label, size in BATCH_OPTIONS:
            if label == self.batch_var.get():
                self.batch_size = size
                return

    # ------------------------------------------------------------- connect
    def _toggle_port(self):
        if self.ser and self.ser.is_open:
            self._close_port()
        else:
            self._open_port()

    def _open_port(self):
        port = self.port_var.get()
        if not port:
            messagebox.showwarning(APP_TITLE, "Select a COM port first.")
            return
        try:
            self.ser = serial.Serial(
                port=port,
                baudrate=int(self.baud_var.get()),
                bytesize=int(self.databits_var.get()),
                parity=self._parity_code(),
                stopbits=self._stopbits_code(),
                timeout=0.2,
            )
            self.ser.rts = self.rts_var.get()
            self.ser.dtr = self.dtr_var.get()
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Could not open {port}:\n{e}")
            self.ser = None
            return

        self.running = True
        self.read_thread = threading.Thread(target=self._read_loop, daemon=True)
        self.read_thread.start()
        self.status_var.set(f"Open ({port} @ {self.baud_var.get()})")
        self.open_btn.configure(text="Close Port")

    def _close_port(self):
        self.running = False
        if self.ser:
            try:
                self.ser.close()
            except Exception:
                pass
        self.ser = None
        self.status_var.set("Closed")
        self.open_btn.configure(text="Open Port")
        self._toggle_autosave_off()

    def _read_loop(self):
        buf = b""
        while self.running and self.ser:
            try:
                chunk = self.ser.read(256)
                if chunk:
                    buf += chunk
                    while b"\n" in buf:
                        line, buf = buf.split(b"\n", 1)
                        text = line.decode(errors="replace")
                        # Strip stray NUL bytes - a common RS-485 half-duplex
                        # artifact (receiver picking up noise as the
                        # transceiver switches direction between frames).
                        # Tkinter/Tcl strings terminate at the first NUL, so
                        # even one of these silently blanks a whole field.
                        text = text.replace("\x00", "")
                        text = text.strip("\r").strip()
                        if text:
                            self.data_queue.put(text)
            except Exception:
                break

    # --------------------------------------------------------------- queue
    def _poll_queue(self):
        drained = 0
        while not self.data_queue.empty() and drained < 1000:
            line = self.data_queue.get_nowait()
            self._handle_line(line)
            drained += 1
        self.root.after(QUEUE_POLL_MS, self._poll_queue)

    def _handle_line(self, line):
        fields = [f.strip() for f in line.split(";")]

        if len(fields) != self.field_count and len(fields) > 0:
            self.field_count = len(fields)
            self.last_known = [""] * self.field_count
            self._build_field_checkboxes()
            self._rebuild_table_columns()

        elapsed = round(self.sample_index * SAMPLE_INTERVAL_S, 2)
        ts = datetime.now()

        # `fields` = exactly what was received, blanks and all - this is
        # what gets saved to file, never altered.
        self.raw_rows.append((ts, line, fields))

        # `display_fields` = same, but blank entries are filled in with the
        # last known non-blank value, for the table and graph only.
        display_fields = carry_forward_fields(fields, self.last_known)

        fault = False
        if display_fields:
            try:
                fault = int(display_fields[-1], 16) != 0
            except ValueError:
                fault = False

        # Queue this row for the next table flush (never dropped - just
        # waiting its turn to be painted). Includes the untouched raw line
        # too, for Hex view mode.
        self.pending_rows.append((elapsed, line, display_fields, fault))
        if len(self.pending_rows) >= self.batch_size:
            self._flush_table()

        self.count_var.set(f"Received: {len(self.raw_rows)}")

        # ---- Graph buffers (full history kept, scaled/capped/wrapped per field) ----
        self.time_buffer.append(elapsed)
        self.sample_index += 1
        for i, f in enumerate(display_fields):
            if i >= len(self.field_buffers):
                break
            val = compute_scaled_value(i, f)
            if val is None:
                # Blank/unparseable field from the device - store as NaN so
                # matplotlib just draws a gap instead of erroring out.
                val = float("nan")
            self.field_buffers[i].append(val)

        # ---- Auto-save ----
        if self.autosave_file:
            self.autosave_file.write(f"{ts.isoformat()}\t{line}\n")
            self.autosave_file.flush()

    def _flush_table(self):
        if not self.pending_rows:
            return
        hex_mode = self.hex_view_var.get()
        for elapsed, raw_line, fields, fault in self.pending_rows:
            if hex_mode:
                values = [raw_line]
            else:
                values = [f"{elapsed:.2f}"] + [format_table_cell(i, f) for i, f in enumerate(fields)]
            tags = ("fault",) if fault else ()
            self.log_tree.insert("", "end", values=values, tags=tags)
        self.pending_rows.clear()

        # Trim on-screen rows only - raw_rows (used for Save) is untouched.
        children = self.log_tree.get_children()
        overflow = len(children) - MAX_TABLE_ROWS
        if overflow > 0:
            for item in children[:overflow]:
                self.log_tree.delete(item)

        if self.autoscroll_var.get():
            children = self.log_tree.get_children()
            if children:
                self.log_tree.see(children[-1])

    # --------------------------------------------------------------- graph
    # Fixed color palette so each parameter's line, axis label, and axis
    # ticks all match - lets you tell which axis belongs to which line at
    # a glance even with several stacked on the right.
    AXIS_COLORS = ["#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd",
                   "#8c564b", "#e377c2", "#7f7f7f", "#bcbd22", "#17becf",
                   "#aec7e8", "#ffbb78"]

    # Canonical field indices for the motor-specific fields (fixed by our
    # FIELD_NAMES layout).
    IDX_KW_M1, IDX_KW_M2 = 2, 3
    IDX_A_M1, IDX_A_M2 = 4, 5

    # Preferred "nice" step sizes for gridlines, smallest to largest.
    # Encoder (degrees) doesn't get anchored to thresholds - it's always
    # lowest priority, per your instruction - it just picks a nice step
    # for its own range independently.
    NICE_STEPS = [0.01, 0.02, 0.025, 0.05, 0.1, 0.2, 0.25, 0.5,
                  1, 2, 2.5, 5, 10, 20, 25, 50, 100]

    @classmethod
    def _nice_step_for_range(cls, span, target_ticks=6):
        if span <= 0:
            return cls.NICE_STEPS[0]
        raw_step = span / target_ticks
        for s in cls.NICE_STEPS:
            if s >= raw_step * 0.7:
                return s
        return cls.NICE_STEPS[-1]

    @classmethod
    def _nice_ticks(cls, ymin, ymax, anchors=None, target_ticks=6):
        """Builds gridline positions at a nice round step, PLUS always
        includes any given anchor values (e.g. active threshold settings)
        even if they don't fall exactly on the regular step - so a
        threshold line is never missed by the grid."""
        step = cls._nice_step_for_range(ymax - ymin, target_ticks)
        ticks = []
        t = (ymin // step) * step
        # guard against float drift producing an extra/missing tick
        while t <= ymax + step * 1e-6:
            ticks.append(round(t, 6))
            t += step
        for a in (anchors or []):
            if ymin - 1e-9 <= a <= ymax + 1e-9 and not any(abs(a - tk) < step * 1e-3 for tk in ticks):
                ticks.append(round(a, 6))
        return sorted(set(ticks))

    @staticmethod
    def _tick_formatter_for_unit(unit):
        if unit == "kW":
            return lambda v, pos: f"{v:.2f}"
        if unit == "A":
            return lambda v, pos: f"{v:.1f}"
        if unit == "\u00b0":
            return lambda v, pos: f"{v:.0f}"
        return lambda v, pos: f"{v:.2f}"

    def _threshold_anchors_by_index(self, checked):
        """Y-values that must land exactly on the grid for each checked
        field - the active threshold settings for that specific field.
        Computed once, before axes exist, using only the checkbox/entry
        state (same source _draw_threshold_lines reads from)."""
        anchors = {i: [] for i in checked}

        for motor, idx in (("M1", self.IDX_KW_M1), ("M2", self.IDX_KW_M2)):
            if idx in anchors and self.belastningsvakt_enabled[motor].get():
                ov = self._parse_number(self.belastningsvakt_open[motor].get())
                cv = self._parse_number(self.belastningsvakt_close[motor].get())
                if ov is not None:
                    anchors[idx].append(ov)
                if cv is not None:
                    anchors[idx].append(cv)

        # Personskydd targets whichever kW axis is present, M1 preferred -
        # same rule _draw_threshold_lines uses.
        pk_idx = self.IDX_KW_M1 if self.IDX_KW_M1 in anchors else (
            self.IDX_KW_M2 if self.IDX_KW_M2 in anchors else None)
        if pk_idx is not None and self.personskydd_var.get():
            val = self._parse_number(self.personskydd_val.get())
            if val is not None:
                anchors[pk_idx].append(val + 0.25)
                anchors[pk_idx].append(val - 0.10)

        for motor, idx in (("M1", self.IDX_A_M1), ("M2", self.IDX_A_M2)):
            if idx in anchors and self.motorskydd_enabled[motor].get():
                ov = self._parse_number(self.motorskydd_open[motor].get())
                cv = self._parse_number(self.motorskydd_close[motor].get())
                if ov is not None:
                    anchors[idx].append(ov * 1.5)
                    anchors[idx].append(ov * 0.5)
                if cv is not None and cv != ov:
                    anchors[idx].append(cv * 1.5)
                    anchors[idx].append(cv * 0.5)

        return anchors

    def _toggle_pause(self):
        self.graph_paused = not self.graph_paused
        self.pause_btn.configure(text="Resume Graph" if self.graph_paused else "Pause Graph")

    def _remove_extra_axes(self):
        for ax in getattr(self, "extra_axes", []):
            try:
                ax.remove()
            except Exception:
                pass
        self.extra_axes = []

    def _clear_graph(self):
        self.time_buffer.clear()
        for buf in self.field_buffers:
            buf.clear()
        self._remove_extra_axes()
        self.ax.clear()
        self.ax.set_xlabel("Time (s)")
        self._user_zoomed = False
        self.canvas.draw_idle()

    def _checked_indices(self):
        return [i for i, var in enumerate(self.field_vars) if var.get()]

    def _checked_field_names(self):
        return [field_label(i) for i in self._checked_indices()]

    def _checked_units(self):
        units = set()
        for name in self._checked_field_names():
            _, unit, _ = field_scale_info(name)
            units.add(unit)
        return units

    def _parse_number(self, text):
        """Accepts either comma or period as the decimal separator."""
        try:
            return float(str(text).strip().replace(",", "."))
        except ValueError:
            return None

    def _update_threshold_visibility(self):
        """Shows/hides each threshold control based on whether its matching
        field is currently ticked in the graph checkboxes."""
        checked = set(self._checked_indices())
        show = {
            "belast_m1": self.IDX_KW_M1 in checked,
            "belast_m2": self.IDX_KW_M2 in checked,
            "personskydd": self.IDX_KW_M1 in checked or self.IDX_KW_M2 in checked,
            "motor_m1": self.IDX_A_M1 in checked,
            "motor_m2": self.IDX_A_M2 in checked,
        }
        ordered = [
            (self.belast_frame_m1, show["belast_m1"]),
            (self.belast_frame_m2, show["belast_m2"]),
            (self.personskydd_frame, show["personskydd"]),
            (self.motorskydd_frame_m1, show["motor_m1"]),
            (self.motorskydd_frame_m2, show["motor_m2"]),
        ]
        for frame, _ in ordered:
            frame.pack_forget()
        for frame, visible in ordered:
            if visible:
                frame.pack(side="left", padx=(0, 16))

    def _compute_direction_runs(self):
        """Returns a list of (start_time, end_time, kind) spans describing
        Encoder 1's direction of travel - 'open' (increasing), 'close'
        (decreasing), or 'reversing' (a quick flip shorter than the
        filter duration, only in that mode). Wrap-aware: a jump across
        the 359->0 boundary is treated as the small real movement it is,
        not a huge fake swing. Spans are sequential and never overlap,
        since each moment in time belongs to exactly one run.

        Direction is judged on a MOVING AVERAGE of the encoder trace, not
        individual samples - real sensor jitter (a count or two of noise)
        can easily flip the sign of a raw sample-to-sample delta even
        during a long, genuinely one-directional sweep. Averaging over a
        short window first actually reduces that noise (comparing two
        raw noisy points directly does NOT - the true signal grows with
        the gap, but so does nothing about the noise, since it's still
        just two individual noisy samples); comparing two *averaged*
        points cuts the noise on each side by roughly sqrt(window)."""
        idx = 0  # Encoder 1
        if idx >= len(self.field_buffers):
            return []
        vals = self.field_buffers[idx]
        times = self.time_buffer
        n = min(len(vals), len(times))
        if n < 2:
            return []

        WINDOW_S = 0.5   # smoothing window - independent of the reversal filter
        window = max(1, round(WINDOW_S / SAMPLE_INTERVAL_S))
        EPS = 0.5        # degrees of NET averaged movement to count as real

        # Trailing moving average (skips NaN gaps rather than poisoning
        # the running sum with them).
        smoothed = [float("nan")] * n
        buf = []
        csum = 0.0
        for k in range(n):
            v = vals[k]
            if v == v:  # not NaN
                buf.append(v)
                csum += v
                if len(buf) > window:
                    csum -= buf.pop(0)
            smoothed[k] = (csum / len(buf)) if buf else float("nan")

        raw_runs = []  # (start_idx, end_idx, direction) direction in {1,-1}
        cur_dir = 0
        cur_start = 0
        for k in range(1, n):
            j = max(0, k - window)
            a, b = smoothed[j], smoothed[k]
            if a != a or b != b:  # NaN (gap in data)
                d = 0
            else:
                delta = b - a
                if delta > 180:
                    delta -= 360
                elif delta < -180:
                    delta += 360
                d = 1 if delta > EPS else (-1 if delta < -EPS else 0)
            if d != cur_dir:
                if cur_dir != 0:
                    raw_runs.append((cur_start, k - 1, cur_dir))
                cur_dir = d
                cur_start = k - 1
        if cur_dir != 0:
            raw_runs.append((cur_start, n - 1, cur_dir))

        filter_s = self._parse_number(self.direction_filter_var.get())
        if filter_s is None or filter_s < 0:
            filter_s = 2.0
        mode = self.direction_mode_var.get()

        results = []
        for s_idx, e_idx, d in raw_runs:
            dur = times[e_idx] - times[s_idx]
            kind = "open" if d > 0 else "close"
            if dur < filter_s:
                if mode == DIRECTION_MODES[2]:
                    kind = "reversing"
                else:
                    continue
            results.append((times[s_idx], times[e_idx], kind))

        # Merge adjacent same-kind spans - in normal mode, a filtered-out
        # blip leaves nothing between two "close" runs that are really one
        # continuous movement; without this they'd render as two spans
        # with a small gap instead of one clean stretch.
        merged = []
        for start, end, kind in results:
            if merged and merged[-1][2] == kind:
                merged[-1] = (merged[-1][0], end, kind)
            else:
                merged.append((start, end, kind))
        return merged

    def _insert_wrap_gaps(self, xs, ys, threshold=180.0):
        """For degree fields, a wrap from e.g. 358deg to 2deg is a small
        real movement but a huge numeric jump - which would otherwise draw
        a near-vertical line across the whole chart. This inserts a NaN
        gap at any jump bigger than `threshold`, so the line breaks
        instead of drawing that misleading chord (and stays fast to
        render, since matplotlib no longer has to draw those extreme
        near-vertical segments)."""
        if len(ys) < 2:
            return xs, ys
        out_x, out_y = [xs[0]], [ys[0]]
        prev = ys[0]
        for x, y in zip(xs[1:], ys[1:]):
            if not (y != y or prev != prev) and abs(y - prev) > threshold:
                out_x.append(x)
                out_y.append(float("nan"))
            out_x.append(x)
            out_y.append(y)
            prev = y
        return out_x, out_y

    def _draw_threshold_lines(self, axis_by_index, color_by_index):
        """Draws optional dashed/dotted threshold lines, per motor and
        direction, on the exact axis (and in the exact color) of the field
        they belong to. Demonstration lines only, not live data.

        M1 and M2 use phase-offset dash patterns (not just different
        colors) - if both land on the exact same value, a plain overlap
        would let whichever is drawn last fully hide the other. Offsetting
        the dash phase means the two colors alternate along the line
        instead, so both are always visible."""

        def kw_axis_for(idx):
            return axis_by_index.get(idx), color_by_index.get(idx)

        # M1 dash pattern starts at phase 0; M2 starts half a cycle later,
        # so if they coincide the colors interleave rather than one
        # covering the other.
        OPEN_STYLE = {"M1": (0, (6, 3)), "M2": (4.5, (6, 3))}
        CLOSE_STYLE = {"M1": (0, (1, 2)), "M2": (1.5, (1, 2))}

        # Belastningsvakt - per motor, open (dashed) / close (dotted)
        for motor, idx in (("M1", self.IDX_KW_M1), ("M2", self.IDX_KW_M2)):
            ax, color = kw_axis_for(idx)
            if ax is None or not self.belastningsvakt_enabled[motor].get():
                continue
            open_v = self._parse_number(self.belastningsvakt_open[motor].get())
            close_v = self._parse_number(self.belastningsvakt_close[motor].get())
            if open_v is not None:
                ax.axhline(open_v, linestyle=OPEN_STYLE[motor], color=color, linewidth=1.6,
                           label=f"Belastningsvakt {motor} \u00d6ppna ({open_v:g} kW)")
            if close_v is not None and close_v != open_v:
                ax.axhline(close_v, linestyle=CLOSE_STYLE[motor], color=color, linewidth=1.6,
                           label=f"Belastningsvakt {motor} St\u00e4ng ({close_v:g} kW)")

        # Personskydd - global, drawn on whichever kW axis is present.
        # Dotted, so it's visually distinct from Belastningsvakt's dashed.
        # Label shows the actual computed upper/lower bounds, not just
        # the base value.
        pk_idx = self.IDX_KW_M1 if self.IDX_KW_M1 in axis_by_index else (
            self.IDX_KW_M2 if self.IDX_KW_M2 in axis_by_index else None)
        if pk_idx is not None and self.personskydd_var.get():
            ax, color = kw_axis_for(pk_idx)
            val = self._parse_number(self.personskydd_val.get())
            if val is not None:
                upper, lower = val + 0.25, val - 0.10
                ax.axhline(upper, linestyle=":", color=color, linewidth=1.6,
                           label=f"Personskydd ({val:g} kW, +{upper:.2f}/-{lower:.2f})")
                ax.axhline(lower, linestyle=":", color=color, linewidth=1.6,
                           label="_nolegend_")

        # Motorskydd - per motor, open (dashed) / close (dotted), +/-50%
        for motor, idx in (("M1", self.IDX_A_M1), ("M2", self.IDX_A_M2)):
            ax, color = kw_axis_for(idx)
            if ax is None or not self.motorskydd_enabled[motor].get():
                continue
            open_v = self._parse_number(self.motorskydd_open[motor].get())
            close_v = self._parse_number(self.motorskydd_close[motor].get())
            if open_v is not None:
                ax.axhline(open_v * 1.5, linestyle=OPEN_STYLE[motor], color=color, linewidth=1.6,
                           label=f"Motorskydd {motor} \u00d6ppna ({open_v:g} A \u00b150%)")
                ax.axhline(open_v * 0.5, linestyle=OPEN_STYLE[motor], color=color, linewidth=1.6, label="_nolegend_")
            if close_v is not None and close_v != open_v:
                ax.axhline(close_v * 1.5, linestyle=CLOSE_STYLE[motor], color=color, linewidth=1.6,
                           label=f"Motorskydd {motor} St\u00e4ng ({close_v:g} A \u00b150%)")
                ax.axhline(close_v * 0.5, linestyle=CLOSE_STYLE[motor], color=color, linewidth=1.6, label="_nolegend_")

    def _draw_graph(self):
        """Draws the current graph state: one color-coded Y-axis per
        checked parameter, a single shared grid at the finest applicable
        step, and threshold lines color-matched to the axis they belong
        to. Always auto-scales to show all data, UNLESS you've actually
        used the zoom/pan toolbar yourself - only then does it keep your
        view across live redraws (until Reset Zoom or the toolbar's home
        button). Does NOT reschedule itself - call directly for an
        immediate redraw (e.g. a checkbox toggled), or via
        _update_graph() on the timer."""
        if self.graph_paused:
            return

        # Only preserve the X-range if you've genuinely zoomed/panned via
        # the toolbar (tracked by _user_zoomed, set only by the
        # xlim_changed callback firing outside our own programmatic
        # updates below) - never just because a previous draw happened to
        # autoscale to some narrow early range.
        preserved_xlim = None
        if getattr(self, "_user_zoomed", False) and not getattr(self, "_zoom_reset_pending", False):
            try:
                preserved_xlim = self.ax.get_xlim()
            except Exception:
                preserved_xlim = None
        self._zoom_reset_pending = False

        self._programmatic_xlim_update = True
        try:
            self._remove_extra_axes()
            self.ax.clear()
            self.ax.set_xlabel("Time (s)")

            checked = self._checked_indices()
            threshold_anchors = self._threshold_anchors_by_index(checked)
            xs = self.time_buffer
            n_total = len(xs)
            step = max(1, n_total // 3000) if n_total else 1
            xs_ds = xs[::step] if n_total else []

            all_handles, all_labels = [], []

            # Movement direction indicator - drawn first (zorder=0) so it
            # sits behind the data lines, not on top of them.
            if self.direction_mode_var.get() != DIRECTION_MODES[0]:
                span_colors = {"open": "#2ecc71", "close": "#e74c3c", "reversing": "#f1c40f"}
                span_labels = {"open": "\u00d6ppnar", "close": "St\u00e4nger", "reversing": "Reversering"}
                seen_kinds = set()
                for start, end, kind in self._compute_direction_runs():
                    lbl = span_labels[kind] if kind not in seen_kinds else "_nolegend_"
                    seen_kinds.add(kind)
                    span = self.ax.axvspan(start, end, color=span_colors[kind], alpha=0.15,
                                            zorder=0, label=lbl)
                    if lbl != "_nolegend_":
                        all_handles.append(span)
                        all_labels.append(lbl)

            axis_by_index = {}
            color_by_index = {}
            outward_offset = 0
            grid_ref_ax = None  # first kW/A axis found - Encoder never wins this, no matter its position

            for pos, i in enumerate(checked):
                # Color is keyed to the field's own fixed index, not its
                # position in the current selection - so a field's color
                # never changes just because you ticked/unticked something
                # else.
                color = self.AXIS_COLORS[i % len(self.AXIS_COLORS)]
                ax = self.ax if pos == 0 else self.ax.twinx()
                if pos > 1:
                    outward_offset += 55
                    ax.spines["right"].set_position(("outward", outward_offset))
                    ax.set_frame_on(True)
                    ax.patch.set_visible(False)
                if pos > 0:
                    self.extra_axes.append(ax)

                axis_by_index[i] = ax
                color_by_index[i] = color

                _, unit, _ = field_scale_info(field_label(i))

                ys = self.field_buffers[i][::step] if i < len(self.field_buffers) else []
                n = min(len(xs_ds), len(ys))
                if n > 0:
                    plot_xs, plot_ys = xs_ds[:n], ys[:n]
                    if unit == "\u00b0":
                        plot_xs, plot_ys = self._insert_wrap_gaps(plot_xs, plot_ys)
                    line, = ax.plot(plot_xs, plot_ys, color=color, label=field_label(i))
                    all_handles.append(line)
                    all_labels.append(field_label(i))

                axis_label = field_label(i) + (f" ({unit})" if unit else "")
                ax.set_ylabel(axis_label, color=color)
                ax.tick_params(axis="y", colors=color)
                if pos > 0:
                    ax.spines["right"].set_color(color)

                # Physical quantities (kW, A, degrees) can't go negative -
                # anchor the bottom at 0 instead of matplotlib's default
                # autoscale padding, which can leave 0 floating awkwardly
                # above the axis floor.
                if unit is not None:
                    ax.set_ylim(bottom=0)

                # Nice round-number gridlines, forced to also include any
                # active threshold value for this field (so a threshold
                # line is never missed by the grid). Encoder gets no
                # anchoring - it's always lowest priority - just its own
                # independent nice step.
                ymin, ymax = ax.get_ylim()
                ax.set_yticks(self._nice_ticks(ymin, ymax, anchors=threshold_anchors.get(i)))
                ax.yaxis.set_major_formatter(FuncFormatter(self._tick_formatter_for_unit(unit)))

                # Encoder (degrees) is always lowest priority for the
                # shared grid - it never becomes the reference here, even
                # if it happens to be the primary axis just because it has
                # the lowest field index. kW/A axes always win.
                if grid_ref_ax is None and unit in ("kW", "A"):
                    grid_ref_ax = ax

            # Saved-test overlays: plotted on the SAME axis as the matching
            # live field, in that field's color but lighter/dashed, so
            # it's clearly a comparison trace and not live data. Only
            # drawn for fields you currently have ticked live - an overlay
            # doesn't add its own axes.
            for overlay in self.overlay_tests:
                for i in checked:
                    ax = axis_by_index.get(i)
                    if ax is None or i >= len(overlay["fields"]):
                        continue
                    ys_o = overlay["fields"][i]
                    xs_o = overlay["time"]
                    n = min(len(xs_o), len(ys_o))
                    if n == 0:
                        continue
                    plot_xs, plot_ys = xs_o[:n], ys_o[:n]
                    _, unit, _ = field_scale_info(field_label(i))
                    if unit == "\u00b0":
                        plot_xs, plot_ys = self._insert_wrap_gaps(plot_xs, plot_ys)
                    label = f"{field_label(i)} (saved: {overlay['name']})"
                    line, = ax.plot(plot_xs, plot_ys, color=color_by_index[i], alpha=0.45,
                                     linestyle="--", linewidth=1.2, label=label)
                    all_handles.append(line)
                    all_labels.append(label)

            # Single shared grid - drawn on whichever axis has a real
            # physical unit (kW or A), NEVER on Encoder, regardless of
            # which axis happens to be primary. Falls back to the primary
            # axis only if nothing with a known unit is being shown at all.
            (grid_ref_ax or self.ax).grid(True, linewidth=0.3, alpha=0.6)

            self._draw_threshold_lines(axis_by_index, color_by_index)

            # Collect threshold line handles/labels from every axis into
            # one combined legend (twin axes don't share legends
            # automatically).
            for ax in [self.ax] + self.extra_axes:
                h, l = ax.get_legend_handles_labels()
                for hh, ll in zip(h, l):
                    if ll != "_nolegend_" and ll not in all_labels:
                        all_handles.append(hh)
                        all_labels.append(ll)

            if all_handles:
                self.ax.legend(all_handles, all_labels, loc="upper left", fontsize=8)

            if preserved_xlim is not None:
                self.ax.set_xlim(preserved_xlim)
        finally:
            self._programmatic_xlim_update = False

        self.canvas.draw_idle()

    def _on_xlim_changed(self, ax):
        """Fires whenever the primary axis's X-range changes. Only treat
        it as a real user zoom/pan if it happened outside our own redraw
        (see the guard flag in _draw_graph)."""
        if not getattr(self, "_programmatic_xlim_update", False):
            self._user_zoomed = True

    def _reset_zoom(self):
        """Forces the next redraw to autoscale the X-axis fully again,
        instead of keeping whatever the user last zoomed/panned to."""
        self._zoom_reset_pending = True
        self._user_zoomed = False
        self._redraw_now()

    def _redraw_now(self):
        """Immediate redraw, e.g. after toggling a checkbox."""
        self._update_threshold_visibility()
        self._draw_graph()

    def _update_graph(self):
        self._draw_graph()
        self.root.after(GRAPH_REFRESH_MS, self._update_graph)

    def _export_graph(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".png",
            filetypes=[("PNG image", "*.png"), ("PDF", "*.pdf"), ("SVG", "*.svg")],
            title="Export Graph")
        if not path:
            return
        try:
            self._write_graph(path)
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Could not export graph:\n{e}")
            return
        messagebox.showinfo(APP_TITLE, f"Graph exported to:\n{path}")

    def _write_graph(self, path):
        self.fig.savefig(path, dpi=150, bbox_inches="tight")

    # ---------------------------------------------------------------- save
    def _clear_all(self):
        self.raw_rows.clear()
        self.pending_rows.clear()
        for item in self.log_tree.get_children():
            self.log_tree.delete(item)
        self.sample_index = 0
        self.count_var.set("Received: 0")
        self._clear_graph()

    def _write_txt(self, path):
        with open(path, "w", encoding="utf-8") as f:
            for ts, line, _ in self.raw_rows:
                f.write(f"{ts.isoformat()}\t{line}\n")

    def _save_txt(self):
        if not self.raw_rows:
            messagebox.showinfo(APP_TITLE, "No data to save yet.")
            return
        path = filedialog.asksaveasfilename(defaultextension=".txt",
                                             filetypes=[("Text file", "*.txt")])
        if not path:
            return
        self._write_txt(path)
        messagebox.showinfo(APP_TITLE, f"Saved {len(self.raw_rows)} rows to:\n{path}")

    def _write_xlsx_raw(self, path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Raw Log"

        max_fields = max(len(fields) for _, _, fields in self.raw_rows)
        header = ["Timestamp", "Raw Line"] + [field_label(i) for i in range(max_fields)]
        ws.append(header)

        for ts, line, fields in self.raw_rows:
            row = [ts.isoformat(), line] + fields + [""] * (max_fields - len(fields))
            ws.append(row)

        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 40
        for i in range(max_fields):
            ws.column_dimensions[get_column_letter(3 + i)].width = 10

        wb.save(path)

    def _save_xlsx(self):
        if not self.raw_rows:
            messagebox.showinfo(APP_TITLE, "No data to save yet.")
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                             filetypes=[("Excel file", "*.xlsx")])
        if not path:
            return
        self._write_xlsx_raw(path)
        messagebox.showinfo(APP_TITLE, f"Saved {len(self.raw_rows)} rows to:\n{path}")

    def _write_xlsx_with_values(self, path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Raw + Values"

        max_fields = max(len(fields) for _, _, fields in self.raw_rows)

        # For each field: a "<name> (hex)" column, and if it has a known
        # unit, a second "<name> (kW/A/deg)" column with the real number.
        header = ["Timestamp", "Raw Line"]
        field_has_unit = []
        for i in range(max_fields):
            name = field_label(i)
            _, unit, _ = field_scale_info(name)
            header.append(f"{name} (hex)")
            field_has_unit.append(unit)
            if unit is not None:
                header.append(f"{name} ({unit})")
        ws.append(header)

        for ts, line, fields in self.raw_rows:
            row = [ts.isoformat(), line]
            for i in range(max_fields):
                hex_str = fields[i] if i < len(fields) else ""
                row.append(hex_str)
                unit = field_has_unit[i]
                if unit is not None:
                    val = compute_scaled_value(i, hex_str)
                    row.append(val if val is not None else "")
            ws.append(row)

        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 40
        for col_idx in range(3, len(header) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12

        wb.save(path)

    def _save_xlsx_with_values(self):
        if not self.raw_rows:
            messagebox.showinfo(APP_TITLE, "No data to save yet.")
            return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                             filetypes=[("Excel file", "*.xlsx")],
                                             title="Save as .xlsx (hex + values)")
        if not path:
            return
        self._write_xlsx_with_values(path)
        messagebox.showinfo(APP_TITLE, f"Saved {len(self.raw_rows)} rows to:\n{path}")

    def _export_all(self):
        if not self.raw_rows:
            messagebox.showinfo(APP_TITLE, "No data to export yet.")
            return
        chosen = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Any", "*.*")],
            title="Export All - choose a base filename (used for all 4 files)")
        if not chosen:
            return

        base, _ext = os.path.splitext(chosen)
        txt_path = base + ".txt"
        xlsx_raw_path = base + "_raw_hex.xlsx"
        xlsx_values_path = base + "_hex_values.xlsx"
        graph_path = base + "_graph.png"

        try:
            self._write_txt(txt_path)
            self._write_xlsx_raw(xlsx_raw_path)
            self._write_xlsx_with_values(xlsx_values_path)
            self._write_graph(graph_path)
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Export All failed partway through:\n{e}")
            return

        names = "\n".join(os.path.basename(p) for p in
                           [txt_path, xlsx_raw_path, xlsx_values_path, graph_path])
        messagebox.showinfo(APP_TITLE, f"Exported 4 files to:\n{os.path.dirname(base) or '.'}\n\n{names}")

    # ------------------------------------------------------------- saved tests
    def _save_current_test(self):
        if not self.raw_rows:
            messagebox.showinfo(APP_TITLE, "No data to save yet.")
            return
        name = simpledialog.askstring(APP_TITLE, "Name this test:", parent=self.root)
        if not name:
            return
        try:
            os.makedirs(SAVED_TESTS_DIR, exist_ok=True)
            safe = "".join(c for c in name if c.isalnum() or c in (" ", "_", "-")).strip() or "test"
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            path = os.path.join(SAVED_TESTS_DIR, f"{safe}_{stamp}.json")
            data = {
                "name": name,
                "timestamp": datetime.now().isoformat(),
                "lines": [line for _, line, _ in self.raw_rows],
            }
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f)
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Could not save test:\n{e}")
            return
        messagebox.showinfo(APP_TITLE, f"Saved test '{name}' ({len(self.raw_rows)} lines).")
        self._refresh_saved_tests_list()

    def _refresh_saved_tests_list(self):
        entries = []
        try:
            os.makedirs(SAVED_TESTS_DIR, exist_ok=True)
            for fn in os.listdir(SAVED_TESTS_DIR):
                if not fn.endswith(".json"):
                    continue
                try:
                    with open(os.path.join(SAVED_TESTS_DIR, fn), "r", encoding="utf-8") as f:
                        d = json.load(f)
                    ts = d.get("timestamp", "")[:16].replace("T", " ")
                    label = f"{d.get('name', fn)}  ({ts})"
                    entries.append((label, fn))
                except Exception:
                    continue
        except Exception:
            pass
        entries.sort()
        self._saved_tests_file_map = {label: fn for label, fn in entries}
        if hasattr(self, "saved_tests_combo"):
            self.saved_tests_combo["values"] = [label for label, _ in entries]

    def _overlay_selected_test(self):
        label = self.saved_tests_var.get()
        fn = getattr(self, "_saved_tests_file_map", {}).get(label)
        if not fn:
            messagebox.showinfo(APP_TITLE, "Pick a saved test from the dropdown first.")
            return
        try:
            with open(os.path.join(SAVED_TESTS_DIR, fn), "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Could not load test:\n{e}")
            return
        lines = data.get("lines", [])
        if not lines:
            messagebox.showinfo(APP_TITLE, "That saved test has no data.")
            return
        time_buf, field_bufs = build_buffers_from_lines(lines, self.field_count)
        self.overlay_tests.append({"name": data.get("name", fn), "time": time_buf, "fields": field_bufs})
        self._redraw_now()

    def _clear_overlays(self):
        self.overlay_tests = []
        self._redraw_now()

    def _import_hex_file(self):
        path = filedialog.askopenfilename(
            title="Import Hex File",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")])
        if not path:
            return
        try:
            with open(path, "r", encoding="utf-8", errors="replace") as f:
                raw_rows = [ln.rstrip("\n").rstrip("\r") for ln in f if ln.strip()]
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Could not read file:\n{e}")
            return

        lines = []
        for raw in raw_rows:
            raw = raw.replace("\x00", "")  # same NUL-strip safety as live reads
            # Our own saved/auto-save format is "timestamp<TAB>line" - a
            # plain hex file (e.g. from Serial Debug Assistant) has no tab,
            # just the line itself.
            if "\t" in raw:
                parts = raw.split("\t", 1)
                line = parts[1] if len(parts) == 2 else raw
            else:
                line = raw
            line = line.strip()
            if line:
                lines.append(line)

        if not lines:
            messagebox.showinfo(APP_TITLE, "No hex lines found in that file.")
            return

        default_name = os.path.splitext(os.path.basename(path))[0]
        name = simpledialog.askstring(APP_TITLE, "Name this imported test:",
                                       initialvalue=default_name, parent=self.root)
        if not name:
            return

        try:
            os.makedirs(SAVED_TESTS_DIR, exist_ok=True)
            safe = "".join(c for c in name if c.isalnum() or c in (" ", "_", "-")).strip() or "imported"
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            out_path = os.path.join(SAVED_TESTS_DIR, f"{safe}_{stamp}.json")
            data = {
                "name": name,
                "timestamp": datetime.now().isoformat(),
                "lines": lines,
            }
            with open(out_path, "w", encoding="utf-8") as f:
                json.dump(data, f)
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Could not save imported test:\n{e}")
            return

        messagebox.showinfo(APP_TITLE, f"Imported '{name}' ({len(lines)} lines) into Saved Tests.")
        self._refresh_saved_tests_list()

    def _toggle_autosave(self):
        if self.autosave_var.get():
            path = filedialog.asksaveasfilename(defaultextension=".txt",
                                                 filetypes=[("Text file", "*.txt")],
                                                 title="Choose auto-save file")
            if not path:
                self.autosave_var.set(False)
                return
            self.autosave_path = path
            self.autosave_file = open(path, "a", encoding="utf-8")
        else:
            self._toggle_autosave_off()

    def _toggle_autosave_off(self):
        if self.autosave_file:
            try:
                self.autosave_file.close()
            except Exception:
                pass
        self.autosave_file = None
        self.autosave_var.set(False)

    # ---------------------------------------------------------------- exit
    def _on_close(self):
        self._toggle_autosave_off()
        self.running = False
        if self.ser:
            try:
                self.ser.close()
            except Exception:
                pass
        self.root.destroy()


def main():
    root = tk.Tk()
    try:
        style = ttk.Style()
        if "vista" in style.theme_names():
            style.theme_use("vista")
    except Exception:
        pass
    app = SerialMonitorApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
